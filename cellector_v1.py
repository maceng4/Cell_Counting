# -*- coding: utf-8 -*-
"""Cellector_v1.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1g0GKs7gaPwTl3z6Z1nUI8uazimH4CpUq

## Import packages
"""

# TensorFlow and tf.keras
import tensorflow as tf
from tensorflow import keras

# Helper libraries
import numpy as np
import random
import matplotlib.pyplot as plt

import skimage as sk
from skimage import io
from skimage import feature
from skimage import filters
from skimage import data
from skimage.feature import blob_dog, blob_log, blob_doh
from skimage.color import rgb2gray

from google.colab import files 
from google.colab import drive 

from PIL import Image
from math import sqrt

"""#WRITE CODE FOR IMAGES TO BE PULLED OFF GITHUB"""





"""## Link with your google drive and load in your image"""

# this links to your google drive 
drive.mount('/drive')

my_path = '/content/drive/My Drive/Cellector/' #you should have a folder in your drive with this name where your files are stored

# PVN IMAGE
#img_pvn1 = Image.open('my_path') #Rutuja use this one
img = sk.io.imread('/drive/My Drive/Cellector/thalamus.png') #I changed this by cutting out Research folder this is for mac

# PVN IMAGE
#np_img_pvn = numpy.array(rgb_image)
image_gray = rgb2gray(img)

plt.imshow(image_gray)
#plt.imshow(image_gray)

"""## Find ALL of the blobs and label training data"""

#SAMPLE IMAGE

# Compute radii in the 3rd column.
#blobs_log[:, 2] = blobs_log[:, 2] * sqrt(2)
#threshold=.1
blobs_log_pvn = blob_log(image_gray, min_sigma=15, max_sigma=50, num_sigma=10, threshold=.1)
blobs_log_pvn[:, 2] = blobs_log_pvn[:, 2] * sqrt(2)

blobs_dog_pvn = blob_dog(image_gray, min_sigma=15, max_sigma=50, threshold=.1)
blobs_dog_pvn[:, 2] = blobs_dog_pvn[:, 2] * sqrt(2)

blobs_doh_pvn = blob_doh(image_gray, min_sigma=15, max_sigma=50, threshold=.1)
blobs_doh_pvn[:, 2] = blobs_doh_pvn[:, 2] * sqrt(2)

blobs_list = [blobs_log_pvn, blobs_dog_pvn, blobs_doh_pvn]
colors = ['lime', 'lime', 'lime']
titles = ['PVN LOG', 'PVN DOG', 'PVN DOH']
sequence = zip(blobs_list, colors, titles)

fig, axes = plt.subplots(1, 3, figsize=(9, 3), sharex=True, sharey=True)
ax = axes.ravel()

for idx, (blobs, color, title) in enumerate(sequence):
    ax[idx].set_title(title)
    ax[idx].imshow(image_gray)
    for blob in blobs:
        y, x, r = blob
        c = plt.Circle((x, y), r, color=color, linewidth=2, fill=False)
        ax[idx].add_patch(c)
    ax[idx].set_axis_off()

plt.tight_layout()
plt.show()

"""# REFINE THIS SO IT'S just blobs dog

# IGNORE IT AND STORE IT (EDGE AND CORNER CASES
"""

len(blobs_dog_pvn)

# This section of code finds all the blobs and puts a 70x70 pixel square around them 
# then stores them in a list 

bounding_box_list = []

for blob in blobs_dog_pvn:
  x, y, r, = blob 
  y = int(y.round(0))
  x = int(x.round(0))
  maxx, maxy = image_gray.shape
  maxx-=1
  maxy-=1 #0-indexing
  left_bound = x-70
  right_bound = x+70
  upper_bound = y+70
  lower_bound = y-70
  if x < 70:
    left_bound = 0
    right_bound = 140-x
  if x > maxx-70:
    right_bound = maxx
    left_bound = 140 - (maxx-x)
  if y < 70:
    lower_bound = 0
    upper_bound = 140-y
  if y > maxy-70:
    lower_bound = maxy
    upper_bound = 140 - (maxy-y)

  bounding_box = image_gray[left_bound:right_bound,lower_bound:upper_bound]
  #Ignore this code. Just a temp solution to the images not being the same size
  #if(len(bounding_box) != 140):
  #  continue;
  #if(len(bounding_box[0]) != 140):
  #  continue;
  bounding_box_list.append(bounding_box)

# let's grab one cell as an example
plt.imshow(bounding_box_list[190])
bounding_box_list[190].shape

# we need a training set, so let's grab random items from the entire blob list 

training_images = random.sample(bounding_box_list,10)

test_images = random.sample(bounding_box_list,10) #can take out the images you used for training by using "not in"

# Commented out IPython magic to ensure Python compatibility.
# now we need to label our training data, press 'c' for cell, or any other key for other.

# matplob inline DOESNT WORK in colaboratory (of course), so this needs to be downloaded and run as ipynb I think... 
# in order to include the user input 

# %matplotlib inline

training_labels = []


def press(event):
    global keypress
    global vals
        
    if event.key == 'c':
        training_labels.append('cell')
    else: 
        training_labels.append('other')
    

    if len(training_labels) == 2:
        fig.canvas.mpl_disconnect(cid)

    return training_labels


for blobss in range(10):
    
    fig = plt.figure()
    plt.imshow(training_images[blobss])
    cid = fig.canvas.mpl_connect('key_press_event', press)

# Commented out IPython magic to ensure Python compatibility.
# SAME PROCESS EXCEPT FOR TESTING DATA

# %matplotlib inline

testing_labels = []


def press(event):
    global keypress
    global vals
        
    if event.key == 'c':
        testing_labels.append('cell')
    else: 
        testing_labels.append('other')
    

    if len(testing_labels) == 2:
        fig.canvas.mpl_disconnect(cid)

    return testing_labels


for blobss in range(10):
    
    fig = plt.figure()
    plt.imshow(test_images[blobss])
    cid = fig.canvas.mpl_connect('key_press_event', press)

for blobss in range(8):
    
    fig = plt.figure()
    plt.imshow(test_images[blobss])

# let's check our labels
training_labels = np.array([2,2,0,0,1,2,0,1,1,2])
testing_labels = np.array([0,2,0,1,1,0,1])

print(len(test_images))               
print(len(training_images))
print(len(training_labels))
print(len(testing_labels))

"""**Accuracy**"""

#Convert data to excel sheet
for blobss in range(8):
    
    fig = plt.figure()
    plt.imshow(test_images[blobss])

pip install xlsxwriter

"""Create an excel sheet:




column 1: image array



column 2: result (cell, not cell)
"""

# import drive 
from google.colab import drive
drive.mount('/content/drive')

from openpyxl import Workbook
wb = Workbook()
test_filename = 'test_workbook.xlsx'

wb.save('/content/drive/My Drive/'+test_filename)

from openpyxl import load_workbook
wb = load_workbook('/content/drive/My Drive/'+test_filename)

import xlwt 
from xlwt import Workbook 

#wb = Workbook() 
  
sheet1 = wb

sheet1.write(1, 0, 'A') 
sheet1.write(2, 0, 'B') 
sheet1.write(3, 0, 'C') 
sheet1.write(4, 0, 'D') 

  
wb.save('test_workbook.xlsx')

"""How to calculate accuracy"""

#Say Data size = 100
#Get the X and Y for all the data

#Step 1: Divide the data into say 10 subsets 

#Step 2:
#test data as first subset
#Remaining data as train data

#Step 3: Find the difference between actual value and the predicted value for each row in test data

#Step 4: this for all subsets and find the accuracy for each subset

X = list(range(1, 100))
n_subsets = 3
##################### Create N Subsets #####################

import more_itertools 

num_rows_data = list(range(1, 100))
list_sub = []
list_sub_begin = []
list_sub_end = []

#divides or groups indices of the data into N subsets
for sub in more_itertools.divide(n_subsets, num_rows_data):
    list_sub = list_sub + [list(sub)]

#Store the divided indices as starting indices of subset and ending indices of subset
for i in range(0, len(list_sub)):
    list_sub_begin = list_sub_begin + [list_sub[i][0]]
    list_sub_end = list_sub_end + [list_sub[i][len(list_sub[i])-1] + 1]

#creating subsets
X_subsets = []
Y_subsets = []  
for i in range(0, n_subsets):
    X_subsets = X_subsets + [X.iloc[list_sub_begin[i]: list_sub_end[i], :]]
    Y_subsets = Y_subsets + [Y[list_sub_begin[i]: list_sub_end[i]]]
  
  ##################### Create N Subsets #####################

################## Splitting train and test data #################
#Considered 1 subset as testing data
#And remaining data as training data

rows = list(range(X_subsets[0], Y_subsets[0]))

test = pd.DataFrame()
train = pd.DataFrame()
for i in range(0, len(mydata)):
    if(i in rows):
      row_test = pd.DataFrame([mydata.iloc[i, :]])
      test = pd.concat([row_test, test], ignore_index=True)
    else:
      row_train = pd.DataFrame([mydata.iloc[i, :]])
      train = pd.concat([row_train, train], ignore_index=True)

#Finding accuracy of testing data prediction
#Go through each test data and subtract the predicted value to the actual value

#ypred is a list of predicted values of rows in test data
#yactual is a list of actual values of rows in test data

#function return the difference between the predicted and actual value
#if differen

def findDiference(ypreds,yactual):
  difference = abs(ypreds-yactual)
  return difference

"""*italicized text*"""

#newt = test_images[3] + np.zeros((71,140),dtype=float) might have to do something like this to fix when bounding boxes are on edges

"""## Begin keras code"""

train_images = np.stack(training_images, axis=0 )
test_images = np.stack(test_images, axis = 0)


train_labels = np.array(training_labels, axis=0)
test_labels = np.array(testing_labels, axis=0)

from google.colab import drive
drive.mount('/content/drive')

training_images[0].shape

class_names = ['cell','bv','other']

# explore data
test_images.shape

# need to set pixel values to a 0-1 scale
train_images = train_images / 255.0

test_images = test_images / 255.0

#Load a previously saved model
#checkpoint_path will be a path to the checkpoint from the current directory, ie 'training_test/cp.ckpt'
def load_weights(model,checkpoint_path):
  model.load_weights(checkpoint_path)

#Build model(if model has not already been built/fitted)
def build_model():
  # Need to build the model, why 128??
  model = keras.Sequential([
      keras.layers.Flatten(input_shape=(140, 140)),
      keras.layers.Dense(128, activation='relu'),
      keras.layers.Dropout(0.5),
      keras.layers.Dense(3)
  ])
  model.compile(optimizer='adam',
              loss=tf.keras.losses.SparseCategoricalCrossentropy(from_logits=True),
              metrics=['accuracy'])
  return model

#Add training data to model
#Params: 
#model: the model to add to
#checkpoint_path: user defined path in case we would like to revert to this version
#training_images: training images
#training_labels: training labels
#epochs: number of epochs to use
def add_to_model(model, checkpoint_path, training_images, training_labels, num_epochs):
  cp_callback = tf.keras.callbacks.ModelCheckpoint(filepath=checkpoint_path,
                                                 save_weights_only=True,
                                                 verbose=1)
  model.fit(training_images, 
          training_labels,  
          epochs=num_epochs,
          callbacks=[cp_callback])

# Proof-of-Concept
 # Need to build the model, why 128??
  model = keras.Sequential([
      keras.layers.Flatten(input_shape=(140, 140)),
      keras.layers.Dense(128, activation='relu'),
      keras.layers.Dropout(0.5),
      keras.layers.Dense(3)
  ])
  model.compile(optimizer='adam',
              loss=tf.keras.losses.SparseCategoricalCrossentropy(from_logits=True),
              metrics=['accuracy'])
  model.fit(training_images, 
          training_labels,  
          epochs=300) #might lead to overfitting if training data is small



# make predictions 

probability_model = tf.keras.Sequential([model, 
                                         tf.keras.layers.Softmax()])

predictions = probability_model.predict(test_images)

predictions[0]

np.argmax(predictions[0]) #try a prediction out [notice how test labels are not involved]

test_labels[0] #try a prediction out [notice how test labels are not involved]

"""## Plot keras results"""

def plot_image(i, predictions_array, true_label, img):
  predictions_array, true_label, img = predictions_array, true_label[i], img[i]
  plt.grid(False)
  plt.xticks([])
  plt.yticks([])

  plt.imshow(img, cmap=plt.cm.binary)

  predicted_label = np.argmax(predictions_array)
  if predicted_label == true_label:
    color = 'blue'
  else:
    color = 'red'

  plt.xlabel("{} {:2.0f}% ({})".format(class_names[predicted_label],
                                100*np.max(predictions_array),
                                class_names[true_label]),
                                color=color)

def plot_value_array(i, predictions_array, true_label):
  predictions_array, true_label = predictions_array, true_label[i]
  plt.grid(False)
  plt.xticks(range(1))
  plt.yticks([])
  thisplot = plt.bar(range(1), predictions_array, color="#777777")
  plt.ylim([0, 1])
  predicted_label = np.argmax(predictions_array)

  thisplot[predicted_label].set_color('red')
  thisplot[true_label].set_color('blue')

#https://www.tensorflow.org/tutorials/keras/classification
# Plot the first X test images, their predicted labels, and the true labels.
# Color correct predictions in blue and incorrect predictions in red.
num_rows = 7
num_cols = 1
num_images = num_rows*num_cols
plt.figure(figsize=(2*2*num_cols, 2*num_rows))
for i in range(num_images):
  plt.subplot(num_rows, 2*num_cols, 2*i+1)
  plot_image(i, predictions[i], test_labels, test_images)
  plt.subplot(num_rows, 2*num_cols, 2*i+2)
  plot_value_array(i, predictions[i], test_labels)
plt.tight_layout()
plt.show()

